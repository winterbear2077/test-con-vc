from __future__ import annotations

import csv
import ipaddress
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Set, Tuple

from nettest.models import NetworkRow, REQUIRED_COLUMNS


def normalize_header(header: Sequence[str]) -> List[str]:
    return [h.strip().lower() for h in header]


def parse_delimited_rows(lines: Iterable[str], delimiter: Optional[str]) -> List[Dict[str, str]]:
    lines = list(lines)
    if not lines:
        return []

    if delimiter is None:
        parts = [p for p in lines[0].strip().split() if p]
        header = normalize_header(parts)
        out: List[Dict[str, str]] = []
        for line_text in lines[1:]:
            if not line_text.strip():
                continue
            values = [p for p in line_text.strip().split() if p]
            if len(values) < len(header):
                values = values + [""] * (len(header) - len(values))
            if len(values) > len(header):
                values = values[: len(header)]
            out.append(dict(zip(header, [v.strip() for v in values])))
        return out

    reader = csv.DictReader(lines, delimiter=delimiter, skipinitialspace=True)
    if not reader.fieldnames:
        return []
    normalized = normalize_header(reader.fieldnames)
    out = []
    for raw in reader:
        row: Dict[str, str] = {}
        for orig_name, norm_name in zip(reader.fieldnames, normalized):
            row[norm_name] = (raw.get(orig_name, "") or "").strip()
        out.append(row)
    return out


def parse_xlsx(path: Path) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as exc:
        raise RuntimeError("xlsx input requires openpyxl. Install: pip install openpyxl") from exc

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = normalize_header([str(x or "").strip() for x in rows[0]])
    out: List[Dict[str, str]] = []
    for values in rows[1:]:
        if not values or not any(v is not None and str(v).strip() for v in values):
            continue
        norm_values = [str(v).strip() if v is not None else "" for v in values]
        if len(norm_values) < len(header):
            norm_values = norm_values + [""] * (len(header) - len(norm_values))
        if len(norm_values) > len(header):
            norm_values = norm_values[: len(header)]
        out.append(dict(zip(header, norm_values)))
    return out


def load_input(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return parse_xlsx(path)

    text = path.read_text(encoding="utf-8")
    lines = text.splitlines()
    if suffix == ".csv":
        return parse_delimited_rows(lines, delimiter=",")
    return parse_delimited_rows(lines, delimiter=None)


def validate_and_normalize(
    raw_rows: List[Dict[str, str]],
    mngt_token: str,
) -> Tuple[List[NetworkRow], List[Dict[str, Any]]]:
    rejected: List[Dict[str, Any]] = []
    accepted: List[NetworkRow] = []
    seen: Set[Tuple[str, str, str, str, str, str]] = set()

    mngt_token_norm = mngt_token.strip().upper()

    for idx, row in enumerate(raw_rows, start=2):
        row_keys = {k.strip().lower() for k in row.keys()}
        if not REQUIRED_COLUMNS.issubset(row_keys):
            rejected.append({"line": idx, "reason": "missing-required-columns", "row": row})
            continue

        vlan = row.get("vlan", "").strip()
        subnet = row.get("subnet", "").strip()
        gw = row.get("gw", "").strip()
        vrf = row.get("vrf", "").strip()
        cluster = row.get("cluster", "").strip()
        datacenter = row.get("datacenter", "").strip()

        if not all([vlan, subnet, gw, vrf, cluster, datacenter]):
            rejected.append({"line": idx, "reason": "empty-required-field", "row": row})
            continue

        try:
            ipaddress.ip_network(subnet, strict=False)
        except ValueError:
            rejected.append({"line": idx, "reason": "invalid-subnet", "row": row})
            continue

        try:
            ipaddress.ip_address(gw)
        except ValueError:
            rejected.append({"line": idx, "reason": "invalid-gateway", "row": row})
            continue

        dedupe_key = (vlan, subnet, gw, vrf, cluster, datacenter)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        cluster_norm = cluster.strip().upper()
        mode = "mngt-esxi" if cluster_norm == mngt_token_norm else "vm-provisioned"
        accepted.append(
            NetworkRow(
                vlan=vlan,
                subnet=subnet,
                gw=gw,
                vrf=vrf,
                cluster=cluster,
                datacenter=datacenter,
                mode=mode,
                source_line=idx,
            )
        )

    return accepted, rejected
