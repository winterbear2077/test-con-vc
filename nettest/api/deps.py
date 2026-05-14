"""Shared state and helper functions used by all API route modules."""
from __future__ import annotations

import csv
import ipaddress
import json
from pathlib import Path
from typing import Dict, List, Tuple

import nettest.db as _db
from nettest.paths import get_workspace, get_static_dir

# ── Paths ─────────────────────────────────────────────────────────────────────
WORKSPACE   = get_workspace()
ARTIFACTS   = WORKSPACE / "artifacts"
CONFIG_FILE = WORKSPACE / "nettest.config.json"
STATIC_DIR  = get_static_dir()

# ── Active run registry: run_id -> {queue, returncode, pid} ───────────────────
_runs: Dict[str, dict] = {}


# ── Config helpers ────────────────────────────────────────────────────────────
def _read_config() -> dict:
    """Read config from SQLite only. Seeded at startup by migrate_config_from_file."""
    return _db.get_config()


def _write_config(data: dict) -> None:
    """Persist config to SQLite only. nettest.config.json is read-only static config."""
    _db.save_config(data)


# ── Input helpers ─────────────────────────────────────────────────────────────
def _input_path() -> Path:
    val = str(_read_config().get("input", "") or "").strip()
    return WORKSPACE / (val if val else "input.csv")


def _parse_input_file(path: Path) -> list:
    """Parse csv/txt/xlsx into list of dicts with normalized keys."""
    rows: list = []
    suffix = path.suffix.lower()
    if suffix in (".csv", ".txt"):
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                rows.append({k.strip().lower(): str(v).strip() for k, v in row.items()})
    elif suffix == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            if ws is None:
                return rows
            headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(max_row=1))]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append({headers[i]: str(v or "").strip() for i, v in enumerate(row)})
        except ImportError:
            pass
    return rows


# ── Row validation ─────────────────────────────────────────────────────────────
_REQUIRED_COLS = ["vlan", "subnet", "gw", "vrf", "cluster", "datacenter"]


def _is_valid_cidr(s: str) -> bool:
    try:
        ipaddress.ip_network(s, strict=False)
        return "/" in s
    except ValueError:
        return False


def _is_valid_ip(s: str) -> bool:
    try:
        ipaddress.ip_address(s)
        return True
    except ValueError:
        return False


def _is_valid_vlan(s: str) -> bool:
    try:
        return 0 <= int(s) <= 4094
    except ValueError:
        return False


def _validate_rows(rows: List[dict]) -> Tuple[List[dict], List[dict]]:
    """Validate parsed rows.

    Returns (valid_rows, rejected_rows).
    Each rejected item: {"line": int, "reason": str, "row": dict}.
    Blank rows (all fields empty) are silently skipped.
    """
    valid: List[dict] = []
    rejected: List[dict] = []

    # Check that the file has all required columns at all
    if rows:
        file_cols = set(rows[0].keys())
        missing_cols = [c for c in _REQUIRED_COLS if c not in file_cols]
        if missing_cols:
            return [], [{"line": 1, "reason": f"File missing required column(s): {', '.join(missing_cols)}", "row": {}}]

    for i, row in enumerate(rows):
        line = i + 2  # 1-indexed + header row

        # Skip entirely blank rows
        if not any(str(v).strip() for v in row.values()):
            continue

        missing = [c for c in _REQUIRED_COLS if not str(row.get(c, "")).strip()]
        if missing:
            rejected.append({"line": line, "reason": f"Missing field(s): {', '.join(missing)}", "row": row})
            continue

        subnet = str(row.get("subnet", "")).strip()
        if not _is_valid_cidr(subnet):
            rejected.append({"line": line, "reason": f"Invalid subnet CIDR: {subnet!r}", "row": row})
            continue

        gw = str(row.get("gw", "")).strip()
        if gw and not _is_valid_ip(gw):
            rejected.append({"line": line, "reason": f"Invalid gateway IP: {gw!r}", "row": row})
            continue

        vlan = str(row.get("vlan", "")).strip()
        if vlan and not _is_valid_vlan(vlan):
            rejected.append({"line": line, "reason": f"Invalid VLAN (must be 0–4094): {vlan!r}", "row": row})
            continue

        valid.append(row)

    return valid, rejected
